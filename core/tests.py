from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .forms import CostItemForm
from .models import Client, CostItem, Product, ProductCostComponent, Quotation, QuotationItem, QuotationItemExtraCost
from .services import recalculate_quotation_item
from .templatetags.pricing_tags import number2


class PricingCrmTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_360ad", verbosity=0)
        cls.admin = User.objects.create_user("admin", password="test-password", is_staff=True)
        cls.regular_user = User.objects.create_user("sales", password="test-password")

    def create_tarpaulin_quote(self):
        quote = Quotation.objects.create(
            client=Client.objects.get(name="Walk-In Customer"),
            customer_type=Quotation.CustomerType.WALK_IN,
            created_by=self.admin,
        )
        product = Product.objects.get(name="Tarpaulin")
        item = QuotationItem.objects.create(
            quotation=quote,
            product=product,
            width=Decimal("1.5"),
            height=Decimal("3"),
            unit=QuotationItem.Unit.FEET,
            quantity=1,
            selling_rate=product.walk_in_rate,
        )
        recalculate_quotation_item(item)
        return quote, item

    def test_admin_only_access(self):
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 302)
        self.client.login(username="sales", password="test-password")
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 302)
        self.client.login(username="admin", password="test-password")
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)

    def test_area_itemized_calculation_and_snapshot(self):
        quote, item = self.create_tarpaulin_quote()
        item.refresh_from_db()
        self.assertEqual(item.selling_total, Decimal("112.50"))
        self.assertEqual(item.cost_breakdown.count(), 7)
        self.assertEqual(item.cost_total, Decimal("38.27"))
        self.assertEqual(quote.subtotal, Decimal("112.50"))
        self.assertEqual(quote.gross_profit, Decimal("74.23"))

    def test_saved_snapshot_does_not_change_with_material_rate(self):
        _, item = self.create_tarpaulin_quote()
        original_cost = item.cost_total
        material = CostItem.objects.get(name="Banner 10oz Tarpaulin")
        material.unit_cost = Decimal("99")
        material.save()
        item.refresh_from_db()
        self.assertEqual(item.cost_total, original_cost)

    def test_job_specific_eyelets_are_itemized(self):
        _, item = self.create_tarpaulin_quote()
        QuotationItemExtraCost.objects.create(
            quotation_item=item,
            cost_item=CostItem.objects.get(name="Eyelet 10mm"),
            basis=ProductCostComponent.Basis.FIXED,
            usage_quantity=4,
        )
        recalculate_quotation_item(item)
        item.refresh_from_db()
        self.assertEqual(item.cost_total, Decimal("40.48"))
        self.assertTrue(item.cost_breakdown.filter(name="Eyelet 10mm", total_cost=Decimal("2.10")).exists())

    def test_piece_product_fixed_and_variable_costs(self):
        quote = Quotation.objects.create(
            client=Client.objects.get(name="Walk-In Customer"),
            customer_type=Quotation.CustomerType.WALK_IN,
            created_by=self.admin,
        )
        product = Product.objects.get(name="Tshirt 1 side A4 DTF Print")
        item = QuotationItem.objects.create(
            quotation=quote,
            product=product,
            quantity=100,
            selling_rate=product.walk_in_rate,
        )
        recalculate_quotation_item(item)
        item.refresh_from_db()
        self.assertEqual(item.cost_total, Decimal("15076.00"))
        self.assertEqual(item.selling_total, Decimal("25000.00"))

    def test_quotation_excel_contains_cost_breakdown(self):
        quote, _ = self.create_tarpaulin_quote()
        self.client.login(username="admin", password="test-password")
        response = self.client.get(reverse("quotation_export_excel", args=[quote.pk]))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Quotation", "Itemized Costs"])
        self.assertEqual(workbook["Itemized Costs"]["B2"].value, "Banner 10oz Tarpaulin")
        self.assertEqual(workbook["Itemized Costs"]["E2"].number_format, "₱#,##0.00")
        self.assertEqual(workbook["Itemized Costs"]["F2"].number_format, "#,##0.00")

    def test_material_crud_create(self):
        self.client.login(username="admin", password="test-password")
        response = self.client.post(
            reverse("cost_item_create"),
            {
                "name": "Test Installation Labor",
                "category": CostItem.Category.INSTALLATION,
                "supplier": "",
                "unit": CostItem.Unit.HOUR,
                "unit_cost": "600",
                "purchase_description": "",
                "notes": "",
                "active": "on",
            },
        )
        self.assertRedirects(response, reverse("cost_item_list"))
        self.assertTrue(CostItem.objects.filter(name="Test Installation Labor").exists())

    def test_admin_pages_render(self):
        quote, item = self.create_tarpaulin_quote()
        self.client.login(username="admin", password="test-password")
        urls = [
            reverse("dashboard"),
            reverse("client_list"),
            reverse("cost_item_list"),
            reverse("product_list"),
            reverse("product_detail", args=[item.product_id]),
            reverse("quotation_list"),
            reverse("quotation_detail", args=[quote.pk]),
        ]
        for url in urls:
            with self.subTest(url=url):
                self.assertEqual(self.client.get(url).status_code, 200)

    def test_master_data_excel_export(self):
        self.client.login(username="admin", password="test-password")
        response = self.client.get(reverse("master_data_export_excel"))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Cost Items", "Products", "Cost Recipes", "Clients"])
        self.assertEqual(workbook["Cost Items"]["E2"].number_format, "₱#,##0.00")
        self.assertEqual(workbook["Cost Recipes"]["E2"].number_format, "#,##0.00")

    def test_numbers_use_commas_and_two_decimal_places(self):
        self.assertEqual(number2(Decimal("1234567.8")), "1,234,567.80")
        product = Product.objects.get(name="Tarpaulin")
        product.walk_in_rate = Decimal("1234567.80")
        product.save()
        self.client.login(username="admin", password="test-password")
        response = self.client.get(reverse("product_detail", args=[product.pk]))
        self.assertContains(response, "₱1,234,567.80")

    def test_decimal_form_inputs_show_two_decimal_places(self):
        material = CostItem.objects.get(name="Banner 10oz Tarpaulin")
        form = CostItemForm(instance=material)
        self.assertIn('value="3.00"', str(form["unit_cost"]))
        self.assertIn('step="0.01"', str(form["unit_cost"]))

    def test_searchable_dropdown_script_is_loaded(self):
        self.client.login(username="admin", password="test-password")
        response = self.client.get(reverse("quotation_create"))
        self.assertRegex(response.content.decode(), r"/static/core/app\.[a-f0-9]+\.js")

    def test_deployment_admin_command(self):
        with patch.dict(
            "os.environ",
            {"ADMIN_USERNAME": "owner", "ADMIN_EMAIL": "owner@example.com", "ADMIN_PASSWORD": "strong-test-password"},
        ):
            call_command("ensure_admin", verbosity=0)
        owner = User.objects.get(username="owner")
        self.assertTrue(owner.is_staff)
        self.assertTrue(owner.is_superuser)
        self.assertTrue(owner.check_password("strong-test-password"))
