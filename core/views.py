from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.db.models import ProtectedError, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import (
    ClientForm,
    CostItemForm,
    ProductCostComponentForm,
    ProductForm,
    QuotationForm,
    QuotationItemExtraCostForm,
    QuotationItemForm,
)
from .models import Client, CostItem, Product, ProductCostComponent, Quotation, QuotationItem, QuotationItemExtraCost
from .services import recalculate_quotation_item


admin_required = user_passes_test(lambda user: user.is_authenticated and user.is_active and user.is_staff, login_url="login")


@admin_required
def dashboard(request):
    context = {
        "client_count": Client.objects.filter(active=True).count(),
        "cost_item_count": CostItem.objects.filter(active=True).count(),
        "product_count": Product.objects.filter(active=True).count(),
        "quotation_count": Quotation.objects.count(),
        "recent_quotations": Quotation.objects.select_related("client", "created_by")[:8],
    }
    return render(request, "core/dashboard.html", context)


def _save_form(request, form_class, template, redirect_name, instance=None, title="", success="Saved successfully."):
    form = form_class(request.POST or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, success)
        return redirect(redirect_name)
    return render(request, template, {"form": form, "title": title})


def _delete_object(request, obj, redirect_name, label):
    if request.method == "POST":
        try:
            obj.delete()
            messages.success(request, f"{label} deleted.")
        except ProtectedError:
            messages.error(request, f"{label} cannot be deleted because it is already used. Mark it inactive instead.")
        return redirect(redirect_name)
    return render(request, "core/confirm_delete.html", {"object": obj, "label": label})


@admin_required
def client_list(request):
    query = request.GET.get("q", "").strip()
    clients = Client.objects.all()
    if query:
        clients = clients.filter(Q(name__icontains=query) | Q(company__icontains=query) | Q(contact_number__icontains=query))
    return render(request, "core/clients.html", {"clients": clients, "query": query})


@admin_required
def client_create(request):
    return _save_form(request, ClientForm, "core/form.html", "client_list", title="Add Client", success="Client added.")


@admin_required
def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)
    return _save_form(request, ClientForm, "core/form.html", "client_list", client, "Edit Client", "Client updated.")


@admin_required
def client_delete(request, pk):
    return _delete_object(request, get_object_or_404(Client, pk=pk), "client_list", "Client")


@admin_required
def cost_item_list(request):
    query = request.GET.get("q", "").strip()
    items = CostItem.objects.all()
    if query:
        items = items.filter(Q(name__icontains=query) | Q(category__icontains=query) | Q(supplier__icontains=query))
    return render(request, "core/materials/list.html", {"items": items, "query": query})


@admin_required
def cost_item_create(request):
    return _save_form(request, CostItemForm, "core/form.html", "cost_item_list", title="Add Cost Item", success="Cost item added.")


@admin_required
def cost_item_edit(request, pk):
    item = get_object_or_404(CostItem, pk=pk)
    return _save_form(request, CostItemForm, "core/form.html", "cost_item_list", item, "Edit Cost Item", "Cost item updated.")


@admin_required
def cost_item_delete(request, pk):
    return _delete_object(request, get_object_or_404(CostItem, pk=pk), "cost_item_list", "Cost item")


@admin_required
def product_list(request):
    query = request.GET.get("q", "").strip()
    products = Product.objects.all()
    if query:
        products = products.filter(name__icontains=query)
    return render(request, "core/products/list.html", {"products": products, "query": query})


@admin_required
def product_create(request):
    form = ProductForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        product = form.save()
        messages.success(request, "Product added. Add its itemized cost recipe next.")
        return redirect("product_detail", pk=product.pk)
    return render(request, "core/form.html", {"form": form, "title": "Add Sellable Product"})


@admin_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=product)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Product updated.")
        return redirect("product_detail", pk=product.pk)
    return render(request, "core/form.html", {"form": form, "title": "Edit Product"})


@admin_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, "core/products/detail.html", {"product": product})


@admin_required
def product_delete(request, pk):
    return _delete_object(request, get_object_or_404(Product, pk=pk), "product_list", "Product")


@admin_required
def component_create(request, product_pk):
    product = get_object_or_404(Product, pk=product_pk)
    form = ProductCostComponentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        component = form.save(commit=False)
        component.product = product
        component.save()
        messages.success(request, "Cost component added.")
        return redirect("product_detail", pk=product.pk)
    return render(request, "core/form.html", {"form": form, "title": f"Add Cost Component – {product.name}"})


@admin_required
def component_edit(request, pk):
    component = get_object_or_404(ProductCostComponent, pk=pk)
    form = ProductCostComponentForm(request.POST or None, instance=component)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Cost component updated.")
        return redirect("product_detail", pk=component.product_id)
    return render(request, "core/form.html", {"form": form, "title": "Edit Cost Component"})


@admin_required
def component_delete(request, pk):
    component = get_object_or_404(ProductCostComponent, pk=pk)
    product_id = component.product_id
    if request.method == "POST":
        component.delete()
        messages.success(request, "Cost component deleted.")
        return redirect("product_detail", pk=product_id)
    return render(request, "core/confirm_delete.html", {"object": component, "label": "Cost component"})


@admin_required
def quotation_list(request):
    query = request.GET.get("q", "").strip()
    quotations = Quotation.objects.select_related("client", "created_by")
    if query:
        quotations = quotations.filter(
            Q(quote_number__icontains=query) | Q(client__name__icontains=query) | Q(client__company__icontains=query)
        )
    return render(request, "core/quotations/list.html", {"quotations": quotations, "query": query})


@admin_required
def quotation_create(request):
    form = QuotationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        quotation = form.save(commit=False)
        quotation.created_by = request.user
        quotation.save()
        messages.success(request, "Quotation created. Add products and services below.")
        return redirect("quotation_detail", pk=quotation.pk)
    return render(request, "core/form.html", {"form": form, "title": "New Quotation"})


@admin_required
def quotation_edit(request, pk):
    quotation = get_object_or_404(Quotation, pk=pk)
    form = QuotationForm(request.POST or None, instance=quotation)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Quotation details updated.")
        return redirect("quotation_detail", pk=quotation.pk)
    return render(request, "core/form.html", {"form": form, "title": f"Edit {quotation.quote_number}"})


@admin_required
def quotation_detail(request, pk):
    quotation = get_object_or_404(
        Quotation.objects.select_related("client", "created_by").prefetch_related(
            "items__product", "items__cost_breakdown", "items__extra_costs__cost_item"
        ),
        pk=pk,
    )
    return render(request, "core/quotations/detail.html", {"quotation": quotation})


@admin_required
def quotation_delete(request, pk):
    return _delete_object(request, get_object_or_404(Quotation, pk=pk), "quotation_list", "Quotation")


@admin_required
def quotation_item_create(request, quote_pk):
    quotation = get_object_or_404(Quotation, pk=quote_pk)
    form = QuotationItemForm(request.POST or None, quotation=quotation)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.quotation = quotation
        item.save()
        recalculate_quotation_item(item)
        messages.success(request, "Quotation item added with itemized cost snapshot.")
        return redirect("quotation_detail", pk=quotation.pk)
    return render(request, "core/form.html", {"form": form, "title": f"Add Item – {quotation.quote_number}"})


@admin_required
def quotation_item_edit(request, pk):
    item = get_object_or_404(QuotationItem.objects.select_related("quotation", "product"), pk=pk)
    form = QuotationItemForm(request.POST or None, instance=item, quotation=item.quotation)
    if request.method == "POST" and form.is_valid():
        item = form.save()
        recalculate_quotation_item(item)
        messages.success(request, "Quotation item and cost snapshot updated.")
        return redirect("quotation_detail", pk=item.quotation_id)
    return render(request, "core/form.html", {"form": form, "title": "Edit Quotation Item"})


@admin_required
def quotation_item_delete(request, pk):
    item = get_object_or_404(QuotationItem, pk=pk)
    quotation_id = item.quotation_id
    if request.method == "POST":
        item.delete()
        messages.success(request, "Quotation item deleted.")
        return redirect("quotation_detail", pk=quotation_id)
    return render(request, "core/confirm_delete.html", {"object": item, "label": "Quotation item"})


@admin_required
def extra_cost_create(request, item_pk):
    item = get_object_or_404(QuotationItem.objects.select_related("quotation", "product"), pk=item_pk)
    form = QuotationItemExtraCostForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        extra = form.save(commit=False)
        extra.quotation_item = item
        extra.save()
        recalculate_quotation_item(item)
        messages.success(request, "Job-specific cost added and totals recalculated.")
        return redirect("quotation_detail", pk=item.quotation_id)
    return render(request, "core/form.html", {"form": form, "title": f"Add Job-Specific Cost – {item.product.name}"})


@admin_required
def extra_cost_edit(request, pk):
    extra = get_object_or_404(QuotationItemExtraCost.objects.select_related("quotation_item__quotation"), pk=pk)
    form = QuotationItemExtraCostForm(request.POST or None, instance=extra)
    if request.method == "POST" and form.is_valid():
        extra = form.save()
        recalculate_quotation_item(extra.quotation_item)
        messages.success(request, "Job-specific cost updated.")
        return redirect("quotation_detail", pk=extra.quotation_item.quotation_id)
    return render(request, "core/form.html", {"form": form, "title": "Edit Job-Specific Cost"})


@admin_required
def extra_cost_delete(request, pk):
    extra = get_object_or_404(QuotationItemExtraCost.objects.select_related("quotation_item__quotation"), pk=pk)
    item = extra.quotation_item
    if request.method == "POST":
        extra.delete()
        recalculate_quotation_item(item)
        messages.success(request, "Job-specific cost deleted.")
        return redirect("quotation_detail", pk=item.quotation_id)
    return render(request, "core/confirm_delete.html", {"object": extra, "label": "Job-specific cost"})


def _excel_response(workbook, filename):
    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _style_sheet(sheet, widths=None):
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        if cell.value is not None:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


@admin_required
def quotation_export_excel(request, pk):
    quotation = get_object_or_404(
        Quotation.objects.select_related("client", "created_by").prefetch_related("items__product", "items__cost_breakdown"),
        pk=pk,
    )
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Quotation"
    summary.append(["360 A.D QUOTATION", quotation.quote_number, "Date", quotation.quotation_date])
    summary.append(["Client", str(quotation.client), "Project", quotation.project_name])
    summary.append(["Customer Type", quotation.get_customer_type_display(), "Status", quotation.get_status_display()])
    summary.append([])
    summary.append(
        ["Product", "Description", "Width", "Height", "Unit", "Qty", "Area/pc", "Rate", "Manual Override", "Cost", "Selling"]
    )
    for item in quotation.items.all():
        summary.append(
            [
                item.product.name,
                item.description,
                float(item.width) if item.width is not None else None,
                float(item.height) if item.height is not None else None,
                item.get_unit_display(),
                float(item.quantity),
                float(item.area_per_piece),
                float(item.selling_rate),
                float(item.selling_price_override) if item.selling_price_override is not None else None,
                float(item.cost_total),
                float(item.selling_total),
            ]
        )
    last_item_row = summary.max_row
    summary.append([])
    first_total_row = summary.max_row + 1
    summary.append(["Total Cost", float(quotation.total_cost)])
    summary.append(["Subtotal", float(quotation.subtotal)])
    summary.append(["Gross Profit", float(quotation.gross_profit)])
    gp_row = summary.max_row + 1
    summary.append(["GP Margin %", float(quotation.gp_margin) / 100])
    vat_rate_row = summary.max_row + 1
    summary.append(["VAT %", float(quotation.vat_percent) / 100])
    summary.append(["VAT Amount", float(quotation.vat_amount)])
    summary.append(["Grand Total", float(quotation.grand_total)])
    last_total_row = summary.max_row
    summary.freeze_panes = "A6"
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 28
    for column in range(3, 12):
        summary.column_dimensions[get_column_letter(column)].width = 14
    dark_fill = PatternFill("solid", fgColor="17324D")
    gold_fill = PatternFill("solid", fgColor="F2B134")
    for cell in summary[1]:
        cell.fill = dark_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in summary[5]:
        cell.fill = dark_fill
        cell.font = Font(color="FFFFFF", bold=True)
    summary.auto_filter.ref = f"A5:K{last_item_row}"
    for row in summary.iter_rows(min_row=6, max_row=last_item_row, min_col=3, max_col=7):
        for cell in row:
            cell.number_format = '#,##0.00'
    for row in summary.iter_rows(min_row=6, max_row=last_item_row, min_col=8, max_col=11):
        for cell in row:
            cell.number_format = '₱#,##0.00'
    for row_number in range(first_total_row, last_total_row + 1):
        summary.cell(row_number, 1).font = Font(bold=True)
        summary.cell(row_number, 2).number_format = '₱#,##0.00'
    summary.cell(last_total_row, 1).fill = gold_fill
    summary.cell(last_total_row, 2).fill = gold_fill
    summary.cell(last_total_row, 1).font = Font(bold=True)
    summary.cell(last_total_row, 2).font = Font(bold=True)
    summary.cell(gp_row, 2).number_format = "0.00%"
    summary.cell(vat_rate_row, 2).number_format = "0.00%"

    costs = workbook.create_sheet("Itemized Costs")
    costs.append(
        [
            "Product",
            "Cost Component",
            "Category",
            "Basis",
            "Unit Cost",
            "Usage",
            "Computed Quantity",
            "Total Cost",
        ]
    )
    for item in quotation.items.all():
        for line in item.cost_breakdown.all():
            costs.append(
                [
                    item.product.name,
                    line.name,
                    line.category,
                    line.basis,
                    float(line.unit_cost),
                    float(line.usage_quantity),
                    float(line.computed_quantity),
                    float(line.total_cost),
                ]
            )
    _style_sheet(costs, [28, 30, 22, 24, 14, 12, 18, 16])
    for row in costs.iter_rows(min_row=2, min_col=5, max_col=8):
        row[0].number_format = '₱#,##0.00'
        row[1].number_format = '#,##0.00'
        row[2].number_format = '#,##0.00'
        row[3].number_format = '₱#,##0.00'

    return _excel_response(workbook, f"{quotation.quote_number}.xlsx")


@admin_required
def master_data_export_excel(request):
    workbook = Workbook()
    materials = workbook.active
    materials.title = "Cost Items"
    materials.append(["Name", "Category", "Supplier", "Unit", "Unit Cost", "VAT Inclusive", "Active", "Notes"])
    for item in CostItem.objects.all():
        materials.append(
            [item.name, item.get_category_display(), item.supplier, item.get_unit_display(), float(item.unit_cost), item.vat_inclusive, item.active, item.notes]
        )
    _style_sheet(materials, [30, 24, 22, 18, 14, 14, 10, 36])
    for cell in materials["E"][1:]:
        cell.number_format = '₱#,##0.00'

    products = workbook.create_sheet("Products")
    products.append(["Product", "Pricing Type", "Basis", "Walk-In Rate", "Tie-Up Rate", "Minimum", "Buffer %", "Active"])
    for product in Product.objects.all():
        products.append(
            [
                product.name,
                product.get_pricing_type_display(),
                product.pricing_basis,
                float(product.walk_in_rate),
                float(product.tie_up_rate),
                float(product.minimum_price),
                float(product.buffer_percent) / 100,
                product.active,
            ]
        )
    _style_sheet(products, [32, 18, 16, 16, 16, 14, 12, 10])
    for row in products.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '₱#,##0.00'
    for cell in products["G"][1:]:
        cell.number_format = "0.00%"

    recipes = workbook.create_sheet("Cost Recipes")
    recipes.append(["Product", "Cost Item", "Category", "Basis", "Usage Quantity", "Unit Cost", "Sequence", "Notes"])
    for component in ProductCostComponent.objects.select_related("product", "cost_item"):
        recipes.append(
            [
                component.product.name,
                component.cost_item.name,
                component.cost_item.get_category_display(),
                component.get_basis_display(),
                float(component.usage_quantity),
                float(component.cost_item.unit_cost),
                component.sequence,
                component.notes,
            ]
        )
    _style_sheet(recipes, [32, 30, 22, 22, 16, 14, 10, 36])
    for cell in recipes["E"][1:]:
        cell.number_format = '#,##0.00'
    for cell in recipes["F"][1:]:
        cell.number_format = '₱#,##0.00'

    clients = workbook.create_sheet("Clients")
    clients.append(["Name", "Company", "Contact", "Email", "Address", "TIN", "Active", "Notes"])
    for client in Client.objects.all():
        clients.append([client.name, client.company, client.contact_number, client.email, client.address, client.tax_id, client.active, client.notes])
    _style_sheet(clients, [24, 28, 18, 28, 38, 18, 10, 36])

    return _excel_response(workbook, "360AD-master-data.xlsx")
